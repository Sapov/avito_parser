from openpyxl import Workbook
import json


def create_exel():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = "Баннера"
    # --шапка таблицы--
    ws.cell(row=1, column=1, value='Заголовок')
    ws.cell(row=1, column=2, value='Описание')
    ws.cell(row=1, column=3, value='link')
    ws.cell(row=1, column=4, value='Price')
    ws.cell(row=1, column=5, value='city')

    with open('items.json') as file:
        st = json.load(file)
        for index, value in enumerate(st):
            print(index, 'Заголовок: ', value['name'])

            ws.cell(row=index + 2, column=1, value=st[index]['name'])
            ws.cell(row=index + 2, column=2, value=st[index]['description'])
            ws.cell(row=index + 2, column=3, value=st[index]['link'])
            ws.cell(row=index + 2, column=4, value=st[index]['price'])
            ws.cell(row=index + 2, column=5, value=st[index]['city'])

    # Save the file
    wb.save("result.xlsx")


if __name__ == '__main__':
    create_exel()